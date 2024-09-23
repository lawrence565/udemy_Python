from openpyxl import load_workbook

wb = load_workbook("Practice/ETL in Python/dodgers.xlsx")
result = []

ws = wb.worksheets[0]
for row in ws.iter_rows(): # 透過 iterate
    # 透過 list comprehension 製作成 List
    result.append([cell.value for cell in row])

sum = 0
for row in result[1:]:
    sum += int(row[11])

print(f"The total number of homerun for Dodger is {sum}")